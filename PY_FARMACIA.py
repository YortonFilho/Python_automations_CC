import pandas as pd
import os
import pyodbc
from functions.data_base import db_connection  # função para conectar com banco de dados
from functions.colors import green, red  # função para utilizar cores nos prints para melhor visualização
from datetime import datetime
from openpyxl import load_workbook

# Variáveis da data para nomear o arquivo a ser salvo
date = datetime.now()
day = f'{date.day:02d}'
month = f'{date.month:02d}'
year = date.year

# Caminho dos arquivos
excel_file = os.path.join('C:/Users/yorton.filho/Downloads/05 - farmacia.xlsx')
output_file = os.path.join(f'C:/Users/yorton.filho/Downloads/19013906000179 - DR_CENTRAL_FARMACIAS {day}{month}{year}.xlsx')

def clean_cpf(cpf):
    # Remove caracteres não numéricos e retorna como string
    if isinstance(cpf, str):
        return ''.join(filter(str.isdigit, cpf))
    return ''

try:
    with db_connection() as connection:  # conectando ao banco de dados
        with connection.cursor() as cursor:  # abrindo central para as querys
            try:
                # executando query para extração de dados
                cursor.execute(""" 
                SELECT
                    DISTINCT UPPER(T.DEPENDENTE) AS NOME,
                    REPLACE(REPLACE(REPLACE(T.CPF_DEPENDENTE, '.', ''), '-', ''), '/', '') AS CPF_DEP
                FROM
                    DADOS_DEPENDENTES_E_TITULARES T
                WHERE
                    T.CPF_DEPENDENTE IS NOT NULL""")
            except Exception as e:
                print(red(f"Erro ao executar comando: {e}"))
                exit()

            # Criando dataFrame para armazenar os dados coletados
            columns = [desc[0] for desc in cursor.description]
            data = cursor.fetchall()
            dataFrame = pd.DataFrame.from_records(data, columns=columns)
            
            # Verificando se o arquivo existe
            if not os.path.exists(excel_file):
                print(red("O arquivo Excel não existe!"))
                exit()

            # abrindo arquivo excel e armazenando as abas que serão verificadas e atualizadas
            try:
                workbook = load_workbook(excel_file)
                sheet1 = workbook.active  # Primeira aba
                sheet2 = workbook.worksheets[1]  # Segunda aba
            except Exception as e:
                print(red(f"Erro ao abrir a planilha existente: {e}"))
                exit()

            # Contando o número de linhas na primeira aba (excluindo cabeçalho)
            num_rows_sheet = sheet1.max_row - 1 
            num_rows_db = len(dataFrame)

            # Validação das linhas apenas na primeira aba
            if num_rows_db < num_rows_sheet:
                print(red("Erro: O número de linhas no banco de dados é menor que o número de linhas na primeira aba da planilha!"))
                exit()
            elif num_rows_db == num_rows_sheet:
                print(red("Erro: O número de linhas no banco de dados é igual ao número de linhas na primeira aba da planilha!"))
                exit()
            elif num_rows_db > num_rows_sheet:
                # Limpar a primeira aba preservando o cabeçalho
                for row in range(2, sheet1.max_row + 1):  # Começando na linha 2
                    for col in range(1, 3):  # Limpa apenas as colunas 1 e 2
                        sheet1.cell(row=row, column=col).value = None

            # Adicionando cabeçalho se necessário
            if sheet1['A1'].value is None:
                sheet1['A1'] = "Nome"
            if sheet1['B1'].value is None:
                sheet1['B1'] = "CPF"

            # Criar uma lista para armazenar os dados formatados
            rows = [(row.NOME, clean_cpf(str(row.CPF_DEP))) for row in dataFrame.itertuples(index=False)]

            # Encontrar a primeira linha vazia na coluna 1
            next_empty_row = 2  # Começando na linha 2
            while sheet1.cell(row=next_empty_row, column=1).value is not None:
                next_empty_row += 1

            # Adicionar os dados ao Excel apenas nas duas primeiras colunas
            for nome, cpf in rows:
                # Insere dados nas colunas 1 e 2
                sheet1.cell(row=next_empty_row, column=1, value=nome)
                sheet1.cell(row=next_empty_row, column=2, value=cpf)
                next_empty_row += 1  # Passa para a próxima linha vazia

            # Salvando as alterações na planilha
            try:
                workbook.save(excel_file)
                print(green("Dados importados com sucesso!"))
            except Exception as e:
                print(red(f"Erro ao salvar as alterações na planilha: {e}"))
                exit()

            # Coletando CPFs da primeira aba
            cpfs_sheet1 = {clean_cpf(sheet1.cell(row=row, column=2).value) for row in range(2, sheet1.max_row + 1)}
            # Para depuração
            # print("CPFs da primeira aba:", cpfs_sheet1)  

            # Coletando CPFs da segunda aba (na primeira coluna)
            cpfs_sheet2 = {clean_cpf(sheet2.cell(row=row, column=1).value) 
                for row in range(2, sheet2.max_row + 1) 
                if isinstance(sheet2.cell(row=row, column=1).value, str)}
             # Para depuração
            # print("CPFs da segunda aba:", cpfs_sheet2) 

            # Identificando CPFs que não estão na segunda aba
            cpfs_nao_encontrados = [(sheet1.cell(row=row, column=1).value, sheet1.cell(row=row, column=2).value)
                for row in range(2, sheet1.max_row + 1) 
                if clean_cpf(sheet1.cell(row=row, column=2).value) not in cpfs_sheet2]

            # Verificando se todos os CPFs foram encontrados
            if not cpfs_nao_encontrados:
                print(red("Erro: Todos os CPFs da primeira aba já foram enviados para a segunda aba!"))
            else:
                # Criando uma nova planilha para CPFs não encontrados
                df_nao_encontrados = pd.DataFrame(cpfs_nao_encontrados, columns=["Nome", "CPF"])
                df_nao_encontrados.to_excel(output_file, index=False)
                print(green(f"Planilha com CPFs não encontrados criada: {output_file}"))

                # Adicionando CPFs não encontrados na segunda aba com a data da modificação
                next_empty_row_sheet2 = sheet2.max_row + 1  # Próxima linha vazia na segunda aba
                for nome, cpf in cpfs_nao_encontrados:
                    sheet2.cell(row=next_empty_row_sheet2, column=1, value=cpf)  # CPF na coluna 1
                    sheet2.cell(row=next_empty_row_sheet2, column=2, value=date.strftime('%d/%m/%Y')) 
                    next_empty_row_sheet2 += 1

                # Salvando as alterações na segunda aba
                try:
                    workbook.save(excel_file)
                    print(green("CPFs não encontrados adicionados à segunda aba com sucesso!"))
                except Exception as e:
                    print(red(f"Erro ao salvar as alterações na segunda aba: {e}"))
                    exit()

except pyodbc.Error as e:
    print(red(f"Erro ao conectar ou interagir com o banco de dados: {e}"))