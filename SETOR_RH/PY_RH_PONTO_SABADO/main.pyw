import pandas as pd
import os
import tkinter as tk
from tkinter import filedialog, messagebox
from datetime import datetime
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from time import sleep

# função para exibir uma janela de erro
def show_error(mensagem):
    root = tk.Tk()
    root.withdraw()
    messagebox.showerror("Erro", mensagem)
    root.destroy()

# função para selecionar os arquivos a serem analisados
def select_files():
    root = tk.Tk()
    root.withdraw()
    file = filedialog.askopenfilenames(filetypes=[("Excel files", "*.xlsx")])
    return file

# cria a janela principal
root = tk.Tk()
root.withdraw()  # oculta a janela principal

# variaveis da data
date = datetime.now()
month = f'{date.month:02d}'
day = f'{date.day:02d}'

# pede pro usuario selecionar as planilhas 
messagebox.showinfo("Selecionar Arquivo", "Selecione a planilha base dos colaboradores!")
collaborator_base = select_files()
if not collaborator_base:
    show_error(f"Nenhum arquivo selecionado. Encerrando programa!")
    exit()

messagebox.showinfo("Selecionar Arquivo", "Selecione todas planilhas de ponto!")
sheets = select_files()
if not sheets:
    show_error(f"Nenhum arquivo selecionado. Encerrando programa!")
    exit()

# variavel para armazenar os pontos dos colaboradores
presence = {}

# analisa cada planilha de ponto
for i, sheet in enumerate(sheets):
    try:
        # ler a planilha
        df = pd.read_excel(sheet)

        # verificar se a primeira coluna (index 0) existe e tentar extrair a data
        if df.shape[1] > 0:
            first_column = str(df.iloc[0, 0])
            data_match = re.search(r'(\d{2}/\d{2}/\d{4})', first_column)

            if data_match:
                data_str = data_match.group(1)
                formatted_date = datetime.strptime(data_str, '%d/%m/%Y').strftime('%d-%m-%Y')
            else:
                formatted_date = 'data_nao_encontrada'

            # verificar se a segunda coluna (index 1) existe
            if df.shape[1] > 1:
                for index, value in enumerate(df.iloc[:, 1]):  # acessa a segunda coluna
                    if isinstance(value, str) and not any(char.isdigit() for char in value):
                        name = value.strip()

                        # verifica as linhas abaixo do nome do colaborador
                        if index + 2 < len(df):
                            linha_baixo_1 = df.iloc[index + 1, 1]
                            linha_baixo_2 = df.iloc[index + 2, 1]

                            if (pd.notna(linha_baixo_1) and linha_baixo_1 != '') and \
                               (pd.notna(linha_baixo_2) and linha_baixo_2 != ''):
                                if name not in presence:
                                    presence[name] = [0] * len(sheets)

                                presence[name][i] = 1
                            else:
                                continue
            else:
                show_error(f"A planilha {sheet} não contém a coluna esperada.")
                exit()
    except Exception as e:
        show_error(f"erro ao processar {sheet}: {e}")
        exit()

# filtrar apenas colaboradores que tem pelo menos uma presença
filtered_presence = {k: v for k, v in presence.items() if any(presence[k])}

# criar um DataFrame a partir do dicionario filtrado
df_names = pd.DataFrame.from_dict(filtered_presence, orient='index', columns=[f'planilha_{i + 1}' for i in range(len(sheet))]).reset_index()
df_names.columns = ['colaborador'] + [datetime.strptime(re.search(r"(\d{2}/\d{2}/\d{4})", str(pd.read_excel(sheet[j]).iloc[0, 0])).group(1), "%d/%m/%Y").strftime("%d-%m-%Y") for j in range(len(sheet))]

# adicionar a coluna total
df_names['total'] = df_names.iloc[:, 1:].sum(axis=1)

# ordenar o DataFrame pelo nome do colaborador
df_names.sort_values(by='colaborador', inplace=True)

# processar a planilha base para obter os cargos
try:
    # ler a planilha base com o cabeçalho na quarta linha
    base_df = pd.read_excel(collaborator_base[0], header=3)
    
    # selecionar as colunas 'I' e 'J'
    col_colaborador = base_df.columns[8]
    col_cargo = base_df.columns[9]  

    # adicionar uma nova coluna para cargos
    list_of_positions = dict(zip(base_df[col_colaborador], base_df[col_cargo]))

    # verificando se todos os colaboradores encontrados têm um cargo
    df_names['cargo'] = df_names['colaborador'].map(list_of_positions).fillna('cargo_nao_encontrado')

    # reordenar as colunas para que 'cargo' fique ao lado de 'colaborador'
    df_names = df_names[['colaborador', 'cargo'] + [col for col in df_names.columns if col not in ['colaborador', 'cargo']]]
    
except Exception as e:
    show_error(f"erro ao processar a planilha base de colaboradores: {e}")
    

# salvar o DataFrame em uma nova planilha na pasta de downloads do usuario
user = os.getlogin()
folder = f'C:/Users/{user}/Downloads/Colaboradores_Ponto_Sábados_{day}_{month}.xlsx'

# formatando a tabela para melhor visualização
with pd.ExcelWriter(folder, engine='openpyxl') as writer:
    df_names.to_excel(writer, index=False, sheet_name='Colaboradores')
    workbook = writer.book
    worksheet = writer.sheets['Colaboradores']

    # aplicar formatação em listras
    for row in range(2, len(df_names) + 2):  # começar da linha 2, pois a linha 1 é o cabeçalho
        if row % 2 == 0:
            for col in range(1, len(df_names.columns) + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

    # formatar cabeçalho
    for col in range(1, len(df_names.columns) + 1):
        header_cell = worksheet.cell(row=1, column=col)
        header_cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")  # cor de fundo preta
        header_cell.font = Font(color="FFFFFF", bold=True)  # cor do texto branca e em negrito

messagebox.showinfo("A lista de colaboradores foi gerada com sucesso!", "O arquivo foi salvo na sua pasta de Downloads!")